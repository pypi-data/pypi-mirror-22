import pandas as pd 
import openpyxl
import os 
from sqlalchemy import create_engine
from sqlalchemy import types
import pymssql
from sistdmyy._version import __version__






#数据库访问控制
class db:
    def __init__(self):
        """
        数据库访问类
        """
        
        self.con_init()
    def con_init(self):
        self.myuser='DmyyReader'
        self.myhost='CLUSTER2014DB2\MSSQLSERVER2'
        self.mypassword="Sjbd*0708"
        self.mydatabase='test_dmyy'

    def frommssql(self,sql):
        con=pymssql.connect(host=self.myhost,user=self.myuser,password=self.mypassword,database=self.mydatabase)
        df=pd.read_sql(sql,con)
        con.close()
        return df

    def tomssql(self,df,name):
        engine=create_engine("mssql+pymssql://%s:%s@%s/%s?charset=utf8"%(self.myuser,self.mypassword,self.myhost,self.mydatabase),encoding='utf-8')
        def sqlcol(dfparam):

            dtypedict = {}
            for i,j in zip(dfparam.columns, dfparam.dtypes):
                
                if "object" in str(j):
                    try:
                        x=int(df[i].str.len().max()/40)+1 
                    except:
                        x=50
                    dtypedict.update({i: types.VARCHAR(length=x*80)})

                if "datetime" in str(j):
                    dtypedict.update({i: stypes.DateTime()})

                if "float" in str(j):
                    dtypedict.update({i: types.Float(precision=3, asdecimal=True)})

                if "int" in str(j):
                    dtypedict.update({i: types.INT()})

            return dtypedict
        outdict=sqlcol(df)
        df.to_sql(name,engine,if_exists='replace',index=False,dtype=outdict)
    def table_exsited(self,name,func=frommssql):
        f=self.frommssql
        sql='select top 1 * from %s'%name
        try:
            df=f(sql)
            return True
        except:
            return False
    def update_sql(self,sql):
        con=pymssql.connect(host=self.myhost,user=self.myuser,password=self.mypassword,database=self.mydatabase)
        cur=con.cursor()
        cur.execute(sql)
        con.commit()
        con.close()

#--------------------------------------------------------A：表输出基本模块-------------------------------------------------------------------------------------------
path1="d:\\UserData\\lanmengfei\\Desktop\\myfile"


#mypath()函数用于自动生成文件路径,递增
def mypath(tail='xlsx',path=path1):
     i=0
     name=path+'\\a1.'+tail
     while os.path.exists(name):
          i+=1
          name=(path+'\\a%d.'+tail)%i
     return name


#yourpath(..),传入表名和文件夹路径
def yourpath(nm,path,tail='xlsx'):
     i=0
     name=path+'\\%s.'%nm+tail
     while os.path.exists(name):
          i+=1
          name=(path+'\\%s(%d).'+tail)%(nm,i)
  
     return name


#输出一个df里的信息，到一张表
def outdf(df,sheet='Sheet1'):
     w=pd.ExcelWriter(mypath('xlsx'))
     df.to_excel(w,sheet_name=sheet,index=False)
     wb=w.book
     ws=w.sheets[sheet]
     fm=wb.add_format({'font_size':'10'})
     ws.set_column('A:AA',8.43,fm)
     w.save()



     
#输出多个df到一个新的excel工作簿的一张sheet上
def outdfs(dfs,sheet='Sheet1',pt=None):
     if pt==None: pt=mypath('xlsx')
     w=pd.ExcelWriter(pt)
     n=0
     wb=w.book
     fm=wb.add_format({'font_size':'10'})

     for df in dfs:
            m=len(df)
            df.to_excel(w,sheet_name=sheet,startrow=n,index=False)
            ws=w.sheets[sheet]
            ws.set_column('A:AA',8.43,fm)
            n+=m+5
     w.save()


#在一张新的工作簿上输出【dfs1,dfs2,dfs3到sheet1 sheet2 sheet3】
def outdfss(dfss,sheets,pt=None):
     if pt==None: pt=mypath('xlsx')
     w=pd.ExcelWriter(pt)
     for i in range(len(dfss)):
          st=sheets[i]
          n=0
          for df in dfss[i]:
               m=len(df)
               df.to_excel(w,sheet_name=st,startrow=n,index=False)
               n+=m+5
               wb=w.book
               ws=w.sheets[st]
               fm=wb.add_format({'font_size':'10'})
               ws.set_column('A:AA',8.43,fm)
               
          print('表-%s完成'%st)
     w.save()
          

     
#找到某张表的‘最后一行’
     
def findtail(pl,sheet='Sheet1'):
     #p1=path+'\\a10.xlsx'
     w=xlrd.open_workbook(pl)
     tb=w.sheet_by_name(sheet)
     a=tb.col(0)
     n=len(a)
     return n
     #if tb.cell(i,0).value=='':print('空')
     #else:print(tb.cell(i,0).value)
     '''
     while i<1000:
         w=0
         if tb.cell(i,1).value!='':
              i+=1
              continue
         if tb.cell(i,1).value=='':
              w=1
              for j in range(10):
                print(i)
                print(tb.cell(i+j,1).value)
                bl=tb.cell(i+j,1).value==''
                w=( w and bl)
         if w==1:break
         else: i+=1
     return i


     '''
#在一张已经存在的表上续写
def adddf(df,file,sheet='Sheet1'):
     if re.search('.xlsx',file)==None:
          file=path+'\\%s.xlsx'%file
     n=findtail(file,sheet)
     book=openpyxl.load_workbook(file)
     w=pd.ExcelWriter(file,engine='openpyxl')
     w.book=book
     w.sheets=dict((ws.title,ws) for ws in book.worksheets)
     df.to_excel(w,sheet,startrow=n+5)
     w.save()
